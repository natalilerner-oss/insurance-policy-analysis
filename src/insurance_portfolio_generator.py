from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from decimal import Decimal
from typing import Dict, List, Tuple
try:
    from .insurance_portfolio_schema import InsurancePortfolioRequest, InsuranceProduct, Coverage
except ImportError:
    from insurance_portfolio_schema import InsurancePortfolioRequest, InsuranceProduct, Coverage

# Hebrew Terminology Reference
INSURANCE_COMPANIES = {
    "harel": "הראל",
    "phoenix": "פניקס",
    "migdal": "מגדל",
    "clal": "כלל",
    "menora": "מנורה",
    "ayalon": "איילון"
}

INSURANCE_PRODUCTS = {
    "health": "בריאות",
    "critical_illness": "מחלות קשות",
    "life": "ביטוח חיים",
    "disability": "אכ\"ע",
    "personal_accident": "תאונות אישיות",
    "service_letter": "כתב שירות"
}

HEALTH_COVERAGES = {
    "surgeries_first_shekel": "ניתוחים שקל ראשון",
    "surgery_alternatives": "טיפולי מחליפי ניתוח",
    "transplants_abroad": "השתלות וטיפולים בחו\"ל",
    "medical_consultation": "ייעוץ רפואי",
    "special_medications": "תרופות מיוחדות",
    "personal_physician": "רופא מלווה אישי",
    "surgeries_abroad": "ביטוח לניתוחים בחו\"ל",
    "ambulatory_services": "שירותיים אמבולטורים",
    "premium_medications": "תרופות פרימיום",
    "fast_diagnosis": "אבחון רפואי מהיר",
    "medical_tech_devices": "טיפולי בטכנולוגיות ואביזר רפואי"
}

# Styling
HEADER_FILL = PatternFill(start_color="4F46E5", fill_type="solid")
HEADER_FONT = Font(bold=True, size=11, name="Arial", color="FFFFFF")
DATA_FONT = Font(size=10, name="Arial")
SUBTOTAL_FILL = PatternFill(start_color="E5E7EB", fill_type="solid")
CURRENCY_FORMAT = '₪#,##0.00'
DATE_FORMAT = 'DD/MM/YYYY'

HEADERS = [
    "שם המבוטח",
    "מס פוליסה",
    "תחילת ביטוח",
    "שם חברת ביטוח",
    "שם המוצר",
    "פירוט",
    "פרמיה",
    "החרגות והנחות"
]

def generate_insurance_portfolio(data: InsurancePortfolioRequest) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "ביטוח"
    ws.sheet_view.rightToLeft = True  # RTL for Hebrew

    # Write headers
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write data
    row = 2
    for product in data.insurance_products:
        row = _write_product(ws, row, product)

    # Grand total
    if row > 2:
        ws.cell(row=row, column=6, value="סה\"כ פרמיה חודשית")
        total_formula = f"=SUM(G2:G{row-1})"
        ws.cell(row=row, column=7, value=total_formula).number_format = CURRENCY_FORMAT
        ws.cell(row=row, column=7).font = Font(bold=True)
        ws.cell(row=row, column=7).fill = SUBTOTAL_FILL

    # Auto-fit columns
    _auto_fit_columns(ws)

    # Conditional formatting for high premiums
    if ws.max_row >= 2:
        ws.conditional_formatting.add(
            f"G2:G{ws.max_row}",
            CellIsRule(operator="greaterThan", formula=["1000"], fill=PatternFill(start_color="FECACA", fill_type="solid"))
        )

    _add_summary_sheet(wb, data)
    _add_comparison_sheet(wb, data)
    _add_projection_sheet(wb, data)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def _write_product(ws, start_row: int, product: InsuranceProduct) -> int:
    """Write a single insurance product, return next row number"""
    row = start_row

    exclusions_discounts = _format_exclusions_discounts(product)

    if product.coverages:
        # Health-type product with coverage breakdown
        for i, coverage in enumerate(product.coverages):
            if i == 0:
                # First row includes all product info
                ws.cell(row=row, column=1, value=product.member_name)
                ws.cell(row=row, column=2, value=product.policy_number)
                ws.cell(row=row, column=3, value=product.start_date).number_format = DATE_FORMAT
                ws.cell(row=row, column=4, value=product.company)
                ws.cell(row=row, column=5, value=product.product_name)

            ws.cell(row=row, column=6, value=coverage.name)
            ws.cell(row=row, column=7, value=float(coverage.premium)).number_format = CURRENCY_FORMAT

            if i == 0 and exclusions_discounts:
                cell = ws.cell(row=row, column=8, value=exclusions_discounts)
                cell.alignment = Alignment(wrap_text=True, vertical="top")

            row += 1

        # Subtotal row
        total = sum(c.premium for c in product.coverages)
        ws.cell(row=row, column=6, value="סה\"כ")
        ws.cell(row=row, column=7, value=float(total)).number_format = CURRENCY_FORMAT
        ws.cell(row=row, column=7).fill = SUBTOTAL_FILL
        ws.cell(row=row, column=7).font = Font(bold=True)
        row += 1
    else:
        # Simple product (life, critical illness, etc.)
        ws.cell(row=row, column=1, value=product.member_name)
        ws.cell(row=row, column=2, value=product.policy_number)
        ws.cell(row=row, column=3, value=product.start_date).number_format = DATE_FORMAT
        ws.cell(row=row, column=4, value=product.company)
        ws.cell(row=row, column=5, value=product.product_name)
        ws.cell(row=row, column=6, value=product.details)
        ws.cell(row=row, column=7, value=float(product.premium or 0)).number_format = CURRENCY_FORMAT

        if exclusions_discounts:
            cell = ws.cell(row=row, column=8, value=exclusions_discounts)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        row += 1

    return row


def _format_exclusions_discounts(product: InsuranceProduct) -> str:
    parts = []
    if product.exclusions:
        parts.append(f"החרגות: {product.exclusions}")
    if product.discounts:
        parts.append(f"הנחות: {product.discounts}")
    return "\n".join(parts)


def _product_premium(product: InsuranceProduct) -> float:
    if product.coverages:
        return float(sum(c.premium for c in product.coverages))
    return float(product.premium or 0)


def _add_summary_sheet(wb: Workbook, data: InsurancePortfolioRequest) -> None:
    ws = wb.create_sheet("סיכום")
    ws.sheet_view.rightToLeft = True

    ws["A1"] = "שם משפחה"
    ws["B1"] = data.family_name
    ws["A2"] = "תאריך דוח"
    ws["B2"] = str(data.report_date)

    total_premium = sum(_product_premium(p) for p in data.insurance_products)
    ws["A3"] = "מספר פוליסות"
    ws["B3"] = len(data.insurance_products)
    ws["A4"] = "סה""כ פרמיה חודשית"
    ws["B4"] = float(total_premium)
    ws["B4"].number_format = CURRENCY_FORMAT

    ws["A6"] = "פרמיה לפי מבוטח"
    ws["A7"] = "מבוטח"
    ws["B7"] = "פרמיה חודשית"
    ws["A7"].font = HEADER_FONT
    ws["B7"].font = HEADER_FONT
    ws["A7"].fill = HEADER_FILL
    ws["B7"].fill = HEADER_FILL

    member_totals: Dict[str, float] = {}
    for product in data.insurance_products:
        member_totals.setdefault(product.member_name, 0)
        member_totals[product.member_name] += _product_premium(product)

    row = 8
    for member, total in member_totals.items():
        ws.cell(row=row, column=1, value=member)
        ws.cell(row=row, column=2, value=float(total)).number_format = CURRENCY_FORMAT
        row += 1

    if member_totals:
        chart = BarChart()
        chart.title = "פרמיה חודשית לפי מבוטח"
        chart.y_axis.title = "פרמיה חודשית"
        chart.x_axis.title = "מבוטח"
        data_ref = Reference(ws, min_col=2, min_row=7, max_row=row - 1)
        cats_ref = Reference(ws, min_col=1, min_row=8, max_row=row - 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws.add_chart(chart, "D7")

    _auto_fit_columns(ws)


def _add_comparison_sheet(wb: Workbook, data: InsurancePortfolioRequest) -> None:
    ws = wb.create_sheet("השוואת פוליסות")
    ws.sheet_view.rightToLeft = True

    policies: List[Tuple[str, InsuranceProduct]] = []
    for idx, product in enumerate(data.insurance_products):
        label = product.policy_number or f"policy_{idx + 1}"
        if product.company:
            label += f" ({product.company})"
        policies.append((label, product))

    # Collect all unique coverage keys (normalized name or type)
    # Map: key -> {policy_idx: premium, name: display_name}
    coverage_map: Dict[str, Dict[str, Any]] = {}
    
    for p_idx, (_, product) in enumerate(policies):
        if product.coverages:
            for coverage in product.coverages:
                # Use name as key, might need better normalization
                key = coverage.name.strip()
                if key not in coverage_map:
                    coverage_map[key] = {"name": key, "premiums": {}}
                
                # If multiple of same type in one policy, sum them? Or assume unique?
                # For now, sum if exists
                existing = coverage_map[key]["premiums"].get(p_idx, 0)
                coverage_map[key]["premiums"][p_idx] = existing + float(coverage.premium)
        else:
             # Product without breakdown (e.g. Life insurance)
             key = product.product_name.strip()
             if key not in coverage_map:
                 coverage_map[key] = {"name": key, "premiums": {}}
             
             existing = coverage_map[key]["premiums"].get(p_idx, 0)
             coverage_map[key]["premiums"][p_idx] = existing + _product_premium(product)

    # Headers
    ws.cell(row=1, column=1, value="כיסוי")
    for col, (label, _) in enumerate(policies, start=2):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Style first col
    ws.cell(row=1, column=1).font = HEADER_FONT
    ws.cell(row=1, column=1).fill = HEADER_FILL

    # Write Data
    row = 2
    for key in sorted(coverage_map.keys()):
        item = coverage_map[key]
        ws.cell(row=row, column=1, value=item["name"])
        
        for p_idx, (label, _) in enumerate(policies):
            col = p_idx + 2
            val = item["premiums"].get(p_idx)
            if val is not None:
                ws.cell(row=row, column=col, value=val).number_format = CURRENCY_FORMAT
            else:
                ws.cell(row=row, column=col, value="-").alignment = Alignment(horizontal='center')
        
        row += 1

    # Total Row
    ws.cell(row=row, column=1, value="סה\"כ").font = Font(bold=True)
    for p_idx in range(len(policies)):
        col = p_idx + 2
        # Sum column
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f"=SUM({col_letter}2:{col_letter}{row-1})").number_format = CURRENCY_FORMAT
        ws.cell(row=row, column=col).font = Font(bold=True)
        ws.cell(row=row, column=col).fill = SUBTOTAL_FILL

    _auto_fit_columns(ws)


def _add_projection_sheet(wb: Workbook, data: InsurancePortfolioRequest) -> None:
    ws = wb.create_sheet("תחזיות פרמיה")
    ws.sheet_view.rightToLeft = True

    ws.cell(row=1, column=1, value="כיסוי")
    ws.cell(row=1, column=2, value="גיל")
    ws.cell(row=1, column=3, value="פרמיה חודשית")
    for col in range(1, 4):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row = 2
    for product in data.insurance_products:
        projections = getattr(product, "premium_projections", None) or []
        for projection_group in projections:
            group_name = projection_group.get("coverage") or projection_group.get("appendix") or product.product_name
            for point in projection_group.get("projections", []) or []:
                ws.cell(row=row, column=1, value=group_name)
                ws.cell(row=row, column=2, value=point.get("age"))
                ws.cell(row=row, column=3, value=point.get("monthly") or 0).number_format = CURRENCY_FORMAT
                row += 1

    if row == 2:
        ws.cell(row=2, column=1, value="אין תחזיות פרמיה זמינות")

    _auto_fit_columns(ws)

def _auto_fit_columns(ws):
    """Auto-fit column widths based on content"""
    for col in range(1, len(HEADERS) + 1):
        max_length = 0
        column_letter = get_column_letter(col)

        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        # Set minimum width
        adjusted_width = max(max_length + 2, 10)
        ws.column_dimensions[column_letter].width = adjusted_width